import csv
import io
import re
from datetime import datetime, timezone
from typing import Any, Literal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.auth.dependencies import CurrentUser, require_tpc_admin
from app.db.supabase import get_supabase_client


ExportFormat = Literal["csv", "xlsx"]
StudentCluster = Literal["placement_ready", "at_risk", "silent_dropout"]
PlacementStatus = Literal["unplaced", "process", "placed"]
InterventionStatus = Literal["pending", "sent", "acknowledged", "completed"]
AlertSeverity = Literal["low", "medium", "high", "critical"]


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_MEDIA_TYPE = "text/csv; charset=utf-8-sig"

STUDENT_COLUMNS = [
    "Full Name",
    "Email",
    "Department",
    "Batch Year",
    "CGPA",
    "Backlogs",
    "Internships",
    "Mock Tests",
    "Avg Mock Score",
    "Certifications",
    "Vigilo Score",
    "Cluster",
    "Placement Probability",
    "Placement Status",
    "Company Placed",
    "Last Login",
    "Resume Updated",
    "Days Inactive",
]
INTERVENTION_COLUMNS = [
    "Student Name",
    "Department",
    "Intervention Type",
    "Status",
    "AI Message Preview",
    "Created At",
    "Sent At",
    "Acknowledged At",
    "Created By",
]
ALERT_COLUMNS = [
    "Student Name",
    "Department",
    "Alert Type",
    "Severity",
    "Message",
    "Triggered At",
    "Is Resolved",
    "Resolved At",
    "Resolved By",
]

SCORE_FILL_RED = PatternFill(fill_type="solid", start_color="FFFFCCCC", end_color="FFFFCCCC")
SCORE_FILL_YELLOW = PatternFill(fill_type="solid", start_color="FFFFF3CC", end_color="FFFFF3CC")
SCORE_FILL_GREEN = PatternFill(fill_type="solid", start_color="FFCCFFCC", end_color="FFCCFFCC")

SEVERITY_FILLS = {
    "critical": PatternFill(fill_type="solid", start_color="FFFFCCCC", end_color="FFFFCCCC"),
    "high": PatternFill(fill_type="solid", start_color="FFFFD8B1", end_color="FFFFD8B1"),
    "medium": PatternFill(fill_type="solid", start_color="FFFFF3CC", end_color="FFFFF3CC"),
    "low": PatternFill(fill_type="solid", start_color="FFCCFFCC", end_color="FFCCFFCC"),
}


router = APIRouter(tags=["exports"])


def _safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _average(values: list[float]) -> float | None:
    if not values:
        return None
    return round(sum(values) / len(values), 2)


def _parse_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo is not None else value.replace(tzinfo=timezone.utc)

    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None

    return parsed if parsed.tzinfo is not None else parsed.replace(tzinfo=timezone.utc)


def _format_datetime(value: Any) -> str | None:
    parsed = _parse_datetime(value)
    if parsed is None:
        return None
    return parsed.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def _days_inactive(last_portal_login: Any) -> int | str:
    parsed = _parse_datetime(last_portal_login)
    if parsed is None:
        return "Never"

    delta = datetime.now(timezone.utc) - parsed.astimezone(timezone.utc)
    return max(delta.days, 0)


def _slugify_filename_part(value: str) -> str:
    sanitized = re.sub(r"[\\/:*?\"<>|]+", "_", value.strip())
    normalized = re.sub(r"\s+", "_", sanitized)
    cleaned = normalized.strip("_")
    return cleaned or "all"


def _export_date() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d")


def _stream_bytes(content: bytes, media_type: str, filename: str) -> StreamingResponse:
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
    }
    return StreamingResponse(io.BytesIO(content), media_type=media_type, headers=headers)


def _csv_bytes(headers: list[str], rows: list[list[Any]]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)

    return buffer.getvalue().encode("utf-8-sig")


def _style_header_row(ws: Worksheet) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _set_students_column_widths(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    for column in "CDEFGHIJKLMNOPQR":
        ws.column_dimensions[column].width = 15


def _apply_score_fills(ws: Worksheet, score_column_index: int) -> None:
    for row_index in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_index, column=score_column_index)
        score = _safe_float(cell.value)
        if score is None:
            continue

        if score < 35:
            cell.fill = SCORE_FILL_RED
        elif score < 65:
            cell.fill = SCORE_FILL_YELLOW
        else:
            cell.fill = SCORE_FILL_GREEN


def _students_xlsx_bytes(rows: list[list[Any]]) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Students"

    ws.append(STUDENT_COLUMNS)
    for row in rows:
        ws.append(row)

    _style_header_row(ws)
    _set_students_column_widths(ws)
    _apply_score_fills(ws, score_column_index=11)

    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def _alerts_xlsx_bytes(rows: list[list[Any]]) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Alerts"

    ws.append(ALERT_COLUMNS)
    for row in rows:
        ws.append(row)

    _style_header_row(ws)
    for column in "ABCDEFGHI":
        ws.column_dimensions[column].width = 20

    severity_column_index = 4
    for row_index in range(2, ws.max_row + 1):
        severity_value = str(ws.cell(row=row_index, column=severity_column_index).value or "").lower()
        fill = SEVERITY_FILLS.get(severity_value)
        if fill is not None:
            ws.cell(row=row_index, column=severity_column_index).fill = fill

    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def _basic_xlsx_bytes(sheet_name: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = sheet_name

    ws.append(headers)
    for row in rows:
        ws.append(row)

    _style_header_row(ws)
    for column in ws.iter_cols(min_col=1, max_col=len(headers), min_row=1, max_row=1):
        column_letter = column[0].column_letter
        ws.column_dimensions[column_letter].width = 20

    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def _fetch_student_bundle(
    department: str | None = None,
    batch_year: int | None = None,
    cluster: StudentCluster | None = None,
    cluster_values: list[str] | None = None,
    placement_status: PlacementStatus | None = None,
) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    client = get_supabase_client()

    profiles_query = (
        client.table("profiles")
        .select("id, full_name, email, department, batch_year")
        .eq("role", "student")
    )
    if department is not None:
        profiles_query = profiles_query.eq("department", department)
    if batch_year is not None:
        profiles_query = profiles_query.eq("batch_year", batch_year)

    profile_rows = profiles_query.execute().data or []
    if not profile_rows:
        return [], {}, {}

    student_ids = [str(row["id"]) for row in profile_rows]

    student_profile_query = (
        client.table("student_profiles")
        .select(
            "id, cgpa, active_backlogs, internship_count, mock_tests_attempted, "
            "mock_avg_score, certifications_count, placement_status, company_placed, "
            "last_portal_login, resume_updated_at"
        )
        .in_("id", student_ids)
    )
    if placement_status is not None:
        student_profile_query = student_profile_query.eq("placement_status", placement_status)

    student_profile_rows = student_profile_query.execute().data or []
    student_profile_map = {str(row["id"]): row for row in student_profile_rows}

    score_query = (
        client.table("vigilo_scores")
        .select("student_id, score, cluster, placement_probability")
        .eq("is_latest", True)
        .in_("student_id", student_ids)
    )
    if cluster is not None:
        score_query = score_query.eq("cluster", cluster)
    elif cluster_values is not None:
        score_query = score_query.in_("cluster", cluster_values)

    score_rows = score_query.execute().data or []
    score_map = {str(row["student_id"]): row for row in score_rows}

    allowed_ids = set(student_ids)
    if placement_status is not None:
        allowed_ids &= set(student_profile_map.keys())
    if cluster is not None or cluster_values is not None:
        allowed_ids &= set(score_map.keys())

    filtered_profiles = [row for row in profile_rows if str(row["id"]) in allowed_ids]
    return filtered_profiles, student_profile_map, score_map


def _build_student_rows(
    profile_rows: list[dict[str, Any]],
    student_profile_map: dict[str, dict[str, Any]],
    score_map: dict[str, dict[str, Any]],
) -> list[list[Any]]:
    rows: list[list[Any]] = []

    for profile in profile_rows:
        student_id = str(profile["id"])
        student_profile = student_profile_map.get(student_id, {})
        score_row = score_map.get(student_id, {})

        row = [
            profile.get("full_name"),
            profile.get("email"),
            profile.get("department"),
            profile.get("batch_year"),
            _safe_float(student_profile.get("cgpa")),
            student_profile.get("active_backlogs"),
            student_profile.get("internship_count"),
            student_profile.get("mock_tests_attempted"),
            _safe_float(student_profile.get("mock_avg_score")),
            student_profile.get("certifications_count"),
            _safe_float(score_row.get("score")),
            score_row.get("cluster"),
            _safe_float(score_row.get("placement_probability")),
            student_profile.get("placement_status"),
            student_profile.get("company_placed"),
            _format_datetime(student_profile.get("last_portal_login")),
            _format_datetime(student_profile.get("resume_updated_at")),
            _days_inactive(student_profile.get("last_portal_login")),
        ]
        rows.append(row)

    rows.sort(key=lambda row: (_safe_float(row[10]) is None, _safe_float(row[10]) or 0.0))
    return rows


def _build_placement_summary_xlsx() -> bytes:
    client = get_supabase_client()

    profile_rows, student_profile_map, score_map = _fetch_student_bundle()
    student_rows = _build_student_rows(profile_rows, student_profile_map, score_map)

    at_risk_profiles, at_risk_student_profiles, at_risk_score_map = _fetch_student_bundle(
        cluster_values=["at_risk", "silent_dropout"]
    )
    at_risk_rows = _build_student_rows(at_risk_profiles, at_risk_student_profiles, at_risk_score_map)

    total_students = len(profile_rows)
    placed_count = sum(
        1
        for row in student_profile_map.values()
        if str(row.get("placement_status") or "") == "placed"
    )
    at_risk_count = sum(1 for row in score_map.values() if str(row.get("cluster") or "") == "at_risk")
    silent_dropout_count = sum(
        1 for row in score_map.values() if str(row.get("cluster") or "") == "silent_dropout"
    )
    placement_rate = round((placed_count / total_students) * 100.0, 2) if total_students > 0 else 0.0
    avg_score = _average([_safe_float(row.get("score")) or 0.0 for row in score_map.values()]) or 0.0

    department_grouped: dict[str, dict[str, Any]] = {}
    for profile in profile_rows:
        student_id = str(profile["id"])
        department = str(profile.get("department") or "Unknown")

        group = department_grouped.setdefault(
            department,
            {
                "total": 0,
                "placed": 0,
                "at_risk": 0,
                "silent_dropout": 0,
                "scores": [],
            },
        )
        group["total"] += 1

        student_profile = student_profile_map.get(student_id, {})
        if str(student_profile.get("placement_status") or "") == "placed":
            group["placed"] += 1

        score_row = score_map.get(student_id, {})
        cluster = str(score_row.get("cluster") or "")
        if cluster == "at_risk":
            group["at_risk"] += 1
        elif cluster == "silent_dropout":
            group["silent_dropout"] += 1

        score = _safe_float(score_row.get("score"))
        if score is not None:
            group["scores"].append(score)

    department_rows: list[list[Any]] = []
    for department in sorted(department_grouped.keys()):
        group = department_grouped[department]
        total = int(group["total"])
        placed = int(group["placed"])
        dept_rate = round((placed / total) * 100.0, 2) if total > 0 else 0.0
        department_rows.append(
            [
                department,
                total,
                placed,
                int(group["at_risk"]),
                int(group["silent_dropout"]),
                dept_rate,
                _average(group["scores"]) or 0.0,
            ]
        )

    drive_rows = client.table("placement_drives").select("company_name, package_lpa").execute().data or []
    packages_by_company: dict[str, list[float]] = {}
    for row in drive_rows:
        company_name = str(row.get("company_name") or "").strip()
        package = _safe_float(row.get("package_lpa"))
        if not company_name or package is None:
            continue
        key = company_name.lower()
        packages_by_company.setdefault(key, []).append(package)

    placement_counts: dict[str, int] = {}
    for row in student_profile_map.values():
        if str(row.get("placement_status") or "") != "placed":
            continue
        company = str(row.get("company_placed") or "").strip()
        if not company:
            continue
        placement_counts[company] = placement_counts.get(company, 0) + 1

    company_rows: list[list[Any]] = []
    for company, count in sorted(placement_counts.items(), key=lambda item: (-item[1], item[0].lower())):
        avg_package = _average(packages_by_company.get(company.lower(), []))
        company_rows.append([company, count, avg_package])

    workbook = Workbook()

    ws_overview = workbook.active
    ws_overview.title = "Overview"
    ws_overview["A1"] = "Placement Summary"
    ws_overview["A1"].font = Font(bold=True)

    overview_rows = [
        ("Total Students", total_students),
        ("Placed", placed_count),
        ("At Risk", at_risk_count),
        ("Silent Dropout", silent_dropout_count),
        ("Placement Rate", placement_rate),
        ("Average Vigilo Score", avg_score),
    ]
    row_index = 3
    for label, value in overview_rows:
        ws_overview.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        ws_overview.cell(row=row_index, column=2, value=value)
        row_index += 1

    ws_overview.column_dimensions["A"].width = 28
    ws_overview.column_dimensions["B"].width = 20

    ws_department = workbook.create_sheet("By Department")
    ws_department.append([
        "Department",
        "Total",
        "Placed",
        "At Risk",
        "Silent Dropout",
        "Placement Rate",
        "Avg Score",
    ])
    for row in department_rows:
        ws_department.append(row)
    _style_header_row(ws_department)
    for column in "ABCDEFG":
        ws_department.column_dimensions[column].width = 18

    ws_company = workbook.create_sheet("By Company")
    ws_company.append(["Company Placed", "Count", "Avg Package"])
    for row in company_rows:
        ws_company.append(row)
    _style_header_row(ws_company)
    ws_company.column_dimensions["A"].width = 28
    ws_company.column_dimensions["B"].width = 12
    ws_company.column_dimensions["C"].width = 14

    ws_risk = workbook.create_sheet("At Risk Students")
    ws_risk.append(STUDENT_COLUMNS)
    for row in at_risk_rows:
        ws_risk.append(row)
    _style_header_row(ws_risk)
    _set_students_column_widths(ws_risk)
    _apply_score_fills(ws_risk, score_column_index=11)

    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


@router.get("/students")
def export_students(
    export_format: ExportFormat = Query(default="xlsx", alias="format"),
    department: str | None = Query(default=None),
    batch_year: int | None = Query(default=None),
    cluster: StudentCluster | None = Query(default=None),
    placement_status: PlacementStatus | None = Query(default=None),
    _: CurrentUser = Depends(require_tpc_admin),
) -> StreamingResponse:
    profile_rows, student_profile_map, score_map = _fetch_student_bundle(
        department=department,
        batch_year=batch_year,
        cluster=cluster,
        placement_status=placement_status,
    )
    rows = _build_student_rows(profile_rows, student_profile_map, score_map)

    dept_part = _slugify_filename_part(department) if department else "all"
    date_part = _export_date()

    if export_format == "csv":
        filename = f"vigilo_students_{dept_part}_{date_part}.csv"
        return _stream_bytes(_csv_bytes(STUDENT_COLUMNS, rows), CSV_MEDIA_TYPE, filename)

    filename = f"vigilo_students_{dept_part}_{date_part}.xlsx"
    return _stream_bytes(_students_xlsx_bytes(rows), XLSX_MEDIA_TYPE, filename)


@router.get("/interventions")
def export_interventions(
    export_format: ExportFormat = Query(default="xlsx", alias="format"),
    from_date: datetime | None = Query(default=None),
    to_date: datetime | None = Query(default=None),
    status_filter: InterventionStatus | None = Query(default=None, alias="status"),
    _: CurrentUser = Depends(require_tpc_admin),
) -> StreamingResponse:
    client = get_supabase_client()

    query = client.table("interventions").select(
        "student_id, intervention_type, status, ai_generated_message, custom_message, "
        "created_at, sent_at, acknowledged_at, created_by"
    )
    if from_date is not None:
        query = query.gte("created_at", from_date.isoformat())
    if to_date is not None:
        query = query.lte("created_at", to_date.isoformat())
    if status_filter is not None:
        query = query.eq("status", status_filter)

    intervention_rows = query.order("created_at", desc=True).execute().data or []

    profile_ids = {
        str(row["student_id"])
        for row in intervention_rows
        if row.get("student_id") is not None
    }
    profile_ids.update(
        {
            str(row["created_by"])
            for row in intervention_rows
            if row.get("created_by") is not None
        }
    )

    profile_map: dict[str, dict[str, Any]] = {}
    if profile_ids:
        profile_rows = (
            client.table("profiles")
            .select("id, full_name, department")
            .in_("id", list(profile_ids))
            .execute()
            .data
            or []
        )
        profile_map = {str(row["id"]): row for row in profile_rows}

    rows: list[list[Any]] = []
    for row in intervention_rows:
        student_id = str(row.get("student_id") or "")
        creator_id = str(row.get("created_by") or "")
        student_profile = profile_map.get(student_id, {})
        creator_profile = profile_map.get(creator_id, {})

        message_value = str(row.get("ai_generated_message") or row.get("custom_message") or "")
        rows.append(
            [
                student_profile.get("full_name"),
                student_profile.get("department"),
                row.get("intervention_type"),
                row.get("status"),
                message_value[:80],
                _format_datetime(row.get("created_at")),
                _format_datetime(row.get("sent_at")),
                _format_datetime(row.get("acknowledged_at")),
                creator_profile.get("full_name") or creator_id,
            ]
        )

    date_part = _export_date()
    if export_format == "csv":
        filename = f"vigilo_interventions_{date_part}.csv"
        return _stream_bytes(_csv_bytes(INTERVENTION_COLUMNS, rows), CSV_MEDIA_TYPE, filename)

    filename = f"vigilo_interventions_{date_part}.xlsx"
    return _stream_bytes(
        _basic_xlsx_bytes("Interventions", INTERVENTION_COLUMNS, rows),
        XLSX_MEDIA_TYPE,
        filename,
    )


@router.get("/alerts")
def export_alerts(
    export_format: ExportFormat = Query(default="xlsx", alias="format"),
    severity: AlertSeverity | None = Query(default=None),
    is_resolved: bool | None = Query(default=None),
    _: CurrentUser = Depends(require_tpc_admin),
) -> StreamingResponse:
    client = get_supabase_client()

    query = client.table("alerts").select(
        "student_id, alert_type, severity, message, triggered_at, is_resolved, resolved_at, resolved_by"
    )
    if severity is not None:
        query = query.eq("severity", severity)
    if is_resolved is not None:
        query = query.eq("is_resolved", is_resolved)

    alert_rows = query.order("triggered_at", desc=True).execute().data or []

    profile_ids = {
        str(row["student_id"])
        for row in alert_rows
        if row.get("student_id") is not None
    }
    profile_ids.update(
        {
            str(row["resolved_by"])
            for row in alert_rows
            if row.get("resolved_by") is not None
        }
    )

    profile_map: dict[str, dict[str, Any]] = {}
    if profile_ids:
        profile_rows = (
            client.table("profiles")
            .select("id, full_name, department")
            .in_("id", list(profile_ids))
            .execute()
            .data
            or []
        )
        profile_map = {str(row["id"]): row for row in profile_rows}

    rows: list[list[Any]] = []
    for row in alert_rows:
        student_id = str(row.get("student_id") or "")
        resolver_id = str(row.get("resolved_by") or "")
        student_profile = profile_map.get(student_id, {})
        resolver_profile = profile_map.get(resolver_id, {})

        rows.append(
            [
                student_profile.get("full_name"),
                student_profile.get("department"),
                row.get("alert_type"),
                row.get("severity"),
                row.get("message"),
                _format_datetime(row.get("triggered_at")),
                row.get("is_resolved"),
                _format_datetime(row.get("resolved_at")),
                resolver_profile.get("full_name") or resolver_id,
            ]
        )

    date_part = _export_date()
    if export_format == "csv":
        filename = f"vigilo_alerts_{date_part}.csv"
        return _stream_bytes(_csv_bytes(ALERT_COLUMNS, rows), CSV_MEDIA_TYPE, filename)

    filename = f"vigilo_alerts_{date_part}.xlsx"
    return _stream_bytes(_alerts_xlsx_bytes(rows), XLSX_MEDIA_TYPE, filename)


@router.get("/placement-summary")
def export_placement_summary(
    _: CurrentUser = Depends(require_tpc_admin),
) -> StreamingResponse:
    date_part = _export_date()
    filename = f"vigilo_placement_summary_{date_part}.xlsx"
    return _stream_bytes(_build_placement_summary_xlsx(), XLSX_MEDIA_TYPE, filename)
